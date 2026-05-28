#!/usr/bin/env python
"""Import latest PT GBR HRIS/manpower Excel files into the MVP HRIS schema."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass, fields, replace
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import text

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.database.models import engine  # noqa: E402
from scripts.hris_data_governance import (  # noqa: E402
    ensure_governance_tables_with_admin,
    record_import_batch,
    sync_data_quality_issues,
)

ASSETS_DIR = PROJECT_ROOT / "assets"
MANPOWER_FILE = ASSETS_DIR / "Data Man Power PT. GBR MARET.xlsx"
STAFF_FILE = ASSETS_DIR / "Data_Staff_Per_20 Maret 2026_PT_GBR .xlsx"

MANPOWER_AS_OF = date(2026, 3, 31)
STAFF_AS_OF = date(2026, 3, 20)
DATE_REVIEW_MARKER = "review_original_join_date="

SOURCE_PRIORITY = {
    "staff_data-tenaga-keluar": 90,
    "staff_data-tenaga-masuk": 80,
    "manpower_tenaga_baru_maret_2026": 70,
    "staff_maret_2026": 65,
    "manpower_lampiran_02_maret_2026": 60,
    "manpower_lampiran_03_maret_2026": 60,
    "manpower_lampiran_01_maret_2026": 50,
}


@dataclass(frozen=True)
class ImportedEmployee:
    source: str
    employee_no: str
    full_name: str
    gender: str | None = None
    birth_place: str | None = None
    birth_date: date | None = None
    join_date: date | None = None
    category_code: str = "PHL"
    employment_type_code: str = "KHL"
    employment_status_code: str = "AKTIF"
    division_name: str = "KANTOR"
    position_name: str = "TENAGA KERJA"
    marital_status_code: str | None = None
    religion_code: str | None = None
    education_code: str | None = None
    ktp_number: str | None = None
    address_text: str | None = None
    bpjs_health_number: str | None = None
    bpjs_employment_number: str | None = None
    notes: str | None = None


def clean(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text_value = " ".join(value.replace("\n", " ").split()).strip()
    else:
        text_value = str(value).strip()
    if not text_value or text_value in {"-", "#VALUE!", "None"}:
        return None
    return text_value


def normalize_code(value: object) -> str | None:
    text_value = clean(value)
    if text_value is None:
        return None
    return re.sub(r"\s+", "_", text_value.upper().replace("/", "_").replace("-", "_"))


def normalize_identity_number(value: object) -> str | None:
    text_value = clean(value)
    if text_value is None:
        return None
    digits = re.sub(r"\D+", "", text_value)
    if 12 <= len(digits) <= 32:
        return digits
    return None


def normalize_card_number(value: object) -> str | None:
    text_value = clean(value)
    if text_value is None:
        return None
    digits = re.sub(r"\D+", "", text_value)
    if 6 <= len(digits) <= 40:
        return digits
    return None


def append_note(notes: str | None, extra: str) -> str:
    base = clean(notes) or ""
    extra_value = clean(extra) or ""
    if not extra_value or extra_value in base:
        return base
    return f"{base} | {extra_value}" if base else extra_value


def source_priority(source: str) -> int:
    return SOURCE_PRIORITY.get(source, 0)


def is_blank(value: object) -> bool:
    return value is None or value == ""


def merge_employee_records(current: ImportedEmployee, incoming: ImportedEmployee) -> ImportedEmployee:
    if source_priority(incoming.source) >= source_priority(current.source):
        base, fallback = incoming, current
    else:
        base, fallback = current, incoming

    values = {}
    for field in fields(ImportedEmployee):
        value = getattr(base, field.name)
        if field.name != "source" and is_blank(value):
            value = getattr(fallback, field.name)
        values[field.name] = value
    return ImportedEmployee(**values)


def dedupe_imported_employees(employees: Iterable[ImportedEmployee]) -> list[ImportedEmployee]:
    merged: dict[tuple[str, str], ImportedEmployee] = {}
    identity_index: dict[str, tuple[str, str]] = {}
    employee_no_index: dict[str, tuple[str, str]] = {}

    for employee in employees:
        key = None
        if employee.ktp_number:
            key = identity_index.get(employee.ktp_number)
        if key is None:
            key = employee_no_index.get(employee.employee_no)
        if key is None:
            key = ("identity", employee.ktp_number) if employee.ktp_number else ("employee_no", employee.employee_no)

        if key in merged:
            merged[key] = merge_employee_records(merged[key], employee)
        else:
            merged[key] = employee

        if employee.ktp_number:
            identity_index[employee.ktp_number] = key
        employee_no_index[employee.employee_no] = key

    return list(merged.values())


def source_snapshot_date(source: str) -> date:
    if source.startswith("staff_"):
        return STAFF_AS_OF
    return MANPOWER_AS_OF


def normalize_imported_employee_dates(employee: ImportedEmployee) -> ImportedEmployee:
    snapshot_date = source_snapshot_date(employee.source)
    if employee.join_date is None or employee.join_date <= snapshot_date:
        return employee
    notes = append_note(employee.notes, employee.source)
    notes = append_note(notes, f"{DATE_REVIEW_MARKER}{employee.join_date.isoformat()}")
    return replace(employee, join_date=snapshot_date, notes=notes)


def normalize_imported_employees(employees: Iterable[ImportedEmployee]) -> list[ImportedEmployee]:
    return [normalize_imported_employee_dates(employee) for employee in employees]


def import_note(employee: ImportedEmployee) -> str:
    return append_note(employee.notes, employee.source)


def effective_date_for_employee(employee: ImportedEmployee) -> date:
    if employee.join_date is not None:
        return employee.join_date
    return STAFF_AS_OF if employee.category_code == "STF" else MANPOWER_AS_OF


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text_value = clean(value)
    if text_value is None:
        return None

    for fmt in (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d-%b-%Y",
        "%Y-%m-%d",
        "%d %B %Y",
        "%d %b %Y",
    ):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue
    return None


def parse_gender(value: object) -> str | None:
    text_value = clean(value)
    if text_value is None:
        return None
    normalized = text_value.upper().replace("-", " ")
    if normalized in {"L", "LAKI", "LAKI LAKI", "PRIA"}:
        return "L"
    if normalized in {"P", "PEREMPUAN", "WANITA"}:
        return "P"
    return None


def map_religion(value: object) -> str | None:
    normalized = normalize_code(value)
    if normalized is None:
        return None
    if normalized in {"KATHOLIK", "CATHOLIC"}:
        return "KATOLIK"
    if normalized in {"KRISTEN_PROTESTAN", "PROTESTAN"}:
        return "KRISTEN"
    return normalized


def map_education(value: object) -> str | None:
    normalized = normalize_code(value)
    if normalized is None:
        return None
    for code in ("S2", "S1", "D3", "SMK", "SMA", "SMP", "SD"):
        if code in normalized:
            return code
    return None


def map_category(value: object, *, staff: bool = False, borongan: bool = False) -> str:
    if staff:
        return "STF"
    if borongan:
        return "BRG"
    normalized = normalize_code(value) or ""
    if "BORONG" in normalized:
        return "BRG"
    if "BULAN" in normalized or normalized == "BLN":
        return "BLN"
    if "SKU" in normalized or "PHT" in normalized or "KHT" in normalized:
        return "SKU"
    return "PHL"


def map_employment_type(value: object, *, staff: bool = False, borongan: bool = False) -> str:
    normalized = normalize_code(value) or ""
    if "PROBATION" in normalized:
        return "PROBATION"
    if "PKWT" in normalized and "PKWTT" not in normalized:
        return "PKWT"
    if "PKWTT" in normalized:
        return "PKWTT"
    if borongan or "BORONG" in normalized:
        return "BORONGAN"
    if "KHL" in normalized or "BHL" in normalized or "PHL" in normalized:
        return "KHL"
    if staff:
        return "PKWTT"
    return "KHL"


def infer_job_family(position_name: str) -> str:
    value = position_name.upper()
    if any(word in value for word in ("PANEN", "POTONG BUAH", "BRONDOL", "BERONDOL", "KUTIP")):
        return "PANEN"
    if "PUPUK" in value:
        return "PUPUK"
    if any(word in value for word in ("RAWAT", "SEMPROT", "BABAT", "CHEMIS")):
        return "RAWAT"
    if any(word in value for word in ("DRIVER", "TRAKSI", "OPERATOR", "HILUX", "DT ", "DUMP")):
        return "TRAKSI"
    if any(word in value for word in ("BENGKEL", "MEKANIK")):
        return "BENGKEL"
    if any(word in value for word in ("SECURITY", "JAGA")):
        return "SECURITY"
    if any(word in value for word in ("SIPIL", "CIVIL", "BANGUN")):
        return "SIPIL"
    return "ADMIN"


def normalize_employee_no(prefix: str, value: object, fallback_number: int) -> str:
    text_value = clean(value)
    if text_value:
        return re.sub(r"\s+", "", text_value.upper())
    return f"{prefix}-{fallback_number:04d}"


def iter_effective_rows(
    worksheet,
    *,
    start_row: int,
    max_col: int,
    no_col: int,
    name_col: int,
    max_scan: int = 1200,
) -> Iterable[tuple[int, tuple[object, ...]]]:
    blanks = 0
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=start_row,
            max_row=start_row + max_scan,
            max_col=max_col,
            values_only=True,
        ),
        start=start_row,
    ):
        no_value = clean(row[no_col - 1] if len(row) >= no_col else None)
        name = clean(row[name_col - 1] if len(row) >= name_col else None)
        if not name and not no_value:
            blanks += 1
            if blanks >= 20:
                break
            continue
        blanks = 0
        if name and name.upper() not in {"NAMA", "NAMA KARYAWAN", "N A M A"}:
            yield row_number, row


def read_staff_employees() -> list[ImportedEmployee]:
    workbook = load_workbook(STAFF_FILE, read_only=True, data_only=True)
    worksheet = workbook["DATA STAFF"]
    employees: list[ImportedEmployee] = []
    for row_number, row in iter_effective_rows(
        worksheet,
        start_row=7,
        max_col=26,
        no_col=2,
        name_col=3,
    ):
        full_name = clean(row[2])
        if not full_name:
            continue
        employee_no = normalize_employee_no("STF", row[3], row_number)
        position_name = clean(row[12]) or "STAFF"
        employee = ImportedEmployee(
            source="staff_maret_2026",
            employee_no=f"STF-{employee_no}",
            full_name=full_name,
            gender=parse_gender(row[8]),
            birth_place=clean(row[4]),
            birth_date=parse_date(row[5]),
            join_date=parse_date(row[9]),
            category_code="STF",
            employment_type_code=map_employment_type(position_name, staff=True),
            employment_status_code="AKTIF",
            division_name="STAFF",
            position_name=position_name,
            marital_status_code=clean(row[13]),
            religion_code=map_religion(row[14]),
            ktp_number=normalize_identity_number(row[3]),
            address_text=clean(row[15]),
            notes=clean(row[25]),
        )
        employees.append(employee)
    workbook.close()
    return employees


def read_staff_movements() -> list[ImportedEmployee]:
    workbook = load_workbook(STAFF_FILE, read_only=True, data_only=True)
    employees: list[ImportedEmployee] = []
    for sheet_name, status_code in (("DATA-TENAGA-MASUK", "AKTIF"), ("DATA-TENAGA-KELUAR", "KELUAR")):
        worksheet = workbook[sheet_name]
        for row_number, row in iter_effective_rows(
            worksheet,
            start_row=8,
            max_col=14,
            no_col=2,
            name_col=3,
            max_scan=80,
        ):
            full_name = clean(row[2])
            if not full_name:
                continue
            position_name = clean(row[5]) or "STAFF"
            join_date = None
            if len(row) >= 9:
                day = clean(row[6])
                month = clean(row[7])
                year = clean(row[8])
                if day and month and year:
                    try:
                        join_date = date(int(float(year)), int(float(month)), int(float(day)))
                    except ValueError:
                        join_date = None
            employees.append(
                ImportedEmployee(
                    source=f"staff_{sheet_name.lower()}",
                    employee_no=f"STF-MOV-{row_number:04d}",
                    full_name=full_name,
                    gender=parse_gender(row[3]),
                    join_date=join_date,
                    category_code="STF",
                    employment_type_code=map_employment_type(position_name, staff=True),
                    employment_status_code=status_code,
                    division_name="STAFF",
                    position_name=position_name,
                    address_text=clean(row[4]),
                )
            )
    workbook.close()
    return employees


def read_manpower_employees() -> list[ImportedEmployee]:
    workbook = load_workbook(MANPOWER_FILE, read_only=True, data_only=True)
    employees: list[ImportedEmployee] = []

    worksheet = workbook["LAMPIRAN 01"]
    for row_number, row in iter_effective_rows(
        worksheet,
        start_row=10,
        max_col=70,
        no_col=2,
        name_col=6,
    ):
        full_name = clean(row[5])
        if not full_name:
            continue
        status_raw = clean(row[11])
        position_name = clean(row[12]) or "TENAGA KERJA"
        employee_no = normalize_employee_no("GBR", row[3], row_number)
        employees.append(
            ImportedEmployee(
                source="manpower_lampiran_01_maret_2026",
                employee_no=employee_no,
                full_name=full_name,
                gender=parse_gender(row[8]),
                birth_place=clean(row[35]) if len(row) >= 36 else None,
                birth_date=parse_date(row[6]),
                join_date=parse_date(row[9]),
                category_code=map_category(status_raw),
                employment_type_code=map_employment_type(status_raw),
                employment_status_code="AKTIF",
                division_name=clean(row[2]) or "KANTOR",
                position_name=position_name,
                marital_status_code=clean(row[13]),
                religion_code=map_religion(row[14]),
                education_code=map_education(row[17]),
                ktp_number=normalize_identity_number(row[4]),
                bpjs_health_number=normalize_card_number(row[57]) if len(row) >= 58 else None,
                bpjs_employment_number=normalize_card_number(row[58]) if len(row) >= 59 else None,
            )
        )

    for sheet_name, start_row, borongan_detail in (
        ("LAMPIRAN 02", 10, "PANEN"),
        ("LAMPIRAN 03", 11, "BRONDOL"),
    ):
        worksheet = workbook[sheet_name]
        for row_number, row in iter_effective_rows(
            worksheet,
            start_row=start_row,
            max_col=22,
            no_col=2,
            name_col=6,
            max_scan=500,
        ):
            full_name = clean(row[5])
            if not full_name:
                continue
            employee_no = normalize_employee_no("BRG", row[3], row_number)
            position_name = clean(row[15]) if len(row) >= 16 else None
            employees.append(
                ImportedEmployee(
                    source=f"manpower_{sheet_name.lower().replace(' ', '_')}_maret_2026",
                    employee_no=employee_no,
                    full_name=full_name,
                    gender=parse_gender(row[6] if sheet_name == "LAMPIRAN 02" else row[7]),
                    birth_date=parse_date(row[11] if sheet_name == "LAMPIRAN 02" else row[12]),
                    join_date=parse_date(row[16] if sheet_name == "LAMPIRAN 02" else row[17]),
                    category_code="BRG",
                    employment_type_code="BORONGAN",
                    employment_status_code="AKTIF",
                    division_name=clean(row[2]) or "BORONGAN",
                    position_name=position_name or f"TENAGA {borongan_detail}",
                    marital_status_code=clean(row[10]) if sheet_name == "LAMPIRAN 02" else None,
                    religion_code=map_religion(row[9] if sheet_name == "LAMPIRAN 02" else row[11]),
                    ktp_number=normalize_identity_number(row[4]),
                    address_text=clean(row[7] if sheet_name == "LAMPIRAN 02" else row[8]),
                    notes=clean(row[18] if sheet_name == "LAMPIRAN 02" else row[19]),
                )
            )

    worksheet = workbook["DATA-TENAGA- KERJA- BARU "]
    for row_number, row in iter_effective_rows(
        worksheet,
        start_row=8,
        max_col=15,
        no_col=2,
        name_col=4,
        max_scan=120,
    ):
        full_name = clean(row[3])
        if not full_name:
            continue
        employee_no = normalize_employee_no("NEW", row[2], row_number)
        active_date = None
        day = clean(row[10]) if len(row) >= 11 else None
        month = clean(row[9]) if len(row) >= 10 else None
        year = clean(row[11]) if len(row) >= 12 else None
        if day and month and year:
            try:
                active_date = date(int(float(year)), int(float(month)), int(float(day)))
            except ValueError:
                active_date = None
        status_raw = clean(row[7])
        employees.append(
            ImportedEmployee(
                source="manpower_tenaga_baru_maret_2026",
                employee_no=employee_no,
                full_name=full_name,
                gender=parse_gender(row[5]),
                join_date=active_date,
                category_code=map_category(status_raw),
                employment_type_code=map_employment_type(status_raw),
                employment_status_code="AKTIF",
                division_name=f"DIVISI {clean(row[8])}" if clean(row[8]) else "KANTOR",
                position_name=status_raw or "TENAGA KERJA",
                ktp_number=normalize_identity_number(row[4]),
                address_text=clean(row[6]),
            )
        )

    workbook.close()
    return employees


def get_reference_id(connection, table: str, id_column: str, code: str) -> int | None:
    row = connection.execute(
        text(f"SELECT {id_column} FROM {table} WHERE code = :code"),
        {"code": code},
    ).first()
    return int(row[0]) if row is not None else None


def get_religion_id(connection, code: str | None) -> int | None:
    if not code:
        return None
    row = connection.execute(text("SELECT id FROM religions WHERE code = :code"), {"code": code}).first()
    return int(row[0]) if row is not None else None


def get_education_id(connection, code: str | None) -> int | None:
    if not code:
        return None
    row = connection.execute(text("SELECT id FROM education_levels WHERE code = :code"), {"code": code}).first()
    return int(row[0]) if row is not None else None


def get_marital_status_id(connection, code: str | None) -> int | None:
    if not code:
        return None
    normalized = code.strip().upper().replace(" ", "")
    row = connection.execute(
        text("SELECT id FROM marital_statuses WHERE code = :code"),
        {"code": normalized},
    ).first()
    return int(row[0]) if row is not None else None


def get_bpjs_type_id(connection, name_pattern: str) -> int | None:
    row = connection.execute(
        text("SELECT bpjs_type_id FROM bpjs_types WHERE upper(bpjs_type_name) LIKE :pattern"),
        {"pattern": f"%{name_pattern.upper()}%"},
    ).first()
    return int(row[0]) if row is not None else None


def ensure_company(connection) -> int:
    row = connection.execute(
        text("SELECT company_id FROM companies WHERE company_code = 'GBR'")
    ).first()
    if row is not None:
        return int(row[0])
    return int(
        connection.execute(
            text(
                """
                INSERT INTO companies (company_code, company_name, company_alias, is_active)
                VALUES ('GBR', 'PT. GERBANG BENUARAYA', 'PT GBR', true)
                RETURNING company_id
                """
            )
        ).scalar_one()
    )


def ensure_estate(connection, company_id: int) -> int:
    row = connection.execute(
        text("SELECT estate_id FROM estates WHERE estate_code = 'BA1'")
    ).first()
    if row is not None:
        return int(row[0])
    return int(
        connection.execute(
            text(
                """
                INSERT INTO estates (company_id, estate_code, estate_name, region, is_active)
                VALUES (:company_id, 'BA1', 'Kebun Batu Ampar', 'Kalimantan Barat', true)
                RETURNING estate_id
                """
            ),
            {"company_id": company_id},
        ).scalar_one()
    )


def ensure_division(connection, estate_id: int, division_name: str) -> int:
    name = clean(division_name) or "KANTOR"
    code = re.sub(r"[^A-Z0-9]+", "_", name.upper()).strip("_")[:32] or "KANTOR"
    row = connection.execute(
        text("SELECT division_id FROM divisions WHERE estate_id = :estate_id AND division_code = :code"),
        {"estate_id": estate_id, "code": code},
    ).first()
    if row is not None:
        return int(row[0])
    return int(
        connection.execute(
            text(
                """
                INSERT INTO divisions (estate_id, division_code, division_name, division_type, is_active)
                VALUES (:estate_id, :code, :name, 'OPERASIONAL', true)
                RETURNING division_id
                """
            ),
            {"estate_id": estate_id, "code": code, "name": name},
        ).scalar_one()
    )


def ensure_position(connection, position_name: str) -> int:
    name = clean(position_name) or "TENAGA KERJA"
    code = re.sub(r"[^A-Z0-9]+", "_", name.upper()).strip("_")[:48] or "TENAGA_KERJA"
    row = connection.execute(
        text("SELECT position_id FROM positions WHERE position_code = :code"),
        {"code": code},
    ).first()
    if row is not None:
        return int(row[0])
    job_family_code = infer_job_family(name)
    job_family_id = get_reference_id(connection, "job_families", "id", job_family_code)
    if job_family_id is None:
        job_family_id = get_reference_id(connection, "job_families", "id", "ADMIN")
    return int(
        connection.execute(
            text(
                """
                INSERT INTO positions (
                    job_family_id, position_code, position_name, level_order, is_staff, is_active
                )
                VALUES (:job_family_id, :code, :name, 0, :is_staff, true)
                RETURNING position_id
                """
            ),
            {
                "job_family_id": job_family_id,
                "code": code,
                "name": name,
                "is_staff": "ASISTEN" in name.upper() or "MANAGER" in name.upper() or "KTU" in name.upper(),
            },
        ).scalar_one()
    )


def find_employee_by_identity(connection, ktp_number: str | None) -> int | None:
    ktp_number = normalize_identity_number(ktp_number)
    if not ktp_number:
        return None
    row = connection.execute(
        text(
            """
            SELECT employee_id
            FROM employee_identities
            WHERE identity_number = :identity_number
              AND upper(identity_type) IN ('NIK', 'KTP')
            LIMIT 1
            """
        ),
        {"identity_number": ktp_number},
    ).first()
    return int(row[0]) if row is not None else None


def upsert_employee(connection, employee: ImportedEmployee) -> int:
    existing = connection.execute(
        text("SELECT employee_id FROM employees WHERE employee_no = :employee_no"),
        {"employee_no": employee.employee_no},
    ).first()
    employee_id = int(existing[0]) if existing is not None else find_employee_by_identity(
        connection,
        employee.ktp_number,
    )

    payload = {
        "employee_no": employee.employee_no,
        "full_name": employee.full_name,
        "gender": employee.gender,
        "birth_place": employee.birth_place,
        "birth_date": employee.birth_date,
        "religion_id": get_religion_id(connection, employee.religion_code),
        "education_id": get_education_id(connection, employee.education_code),
        "marital_status_id": get_marital_status_id(connection, employee.marital_status_code),
        "is_active": employee.employment_status_code != "KELUAR",
    }
    if employee_id is None:
        return int(
            connection.execute(
                text(
                    """
                    INSERT INTO employees (
                        employee_no, full_name, gender, birth_place, birth_date,
                        religion_id, education_id, marital_status_id, is_active
                    )
                    VALUES (
                        :employee_no, :full_name, :gender, :birth_place, :birth_date,
                        :religion_id, :education_id, :marital_status_id, :is_active
                    )
                    RETURNING employee_id
                    """
                ),
                payload,
            ).scalar_one()
        )

    connection.execute(
        text(
            """
            UPDATE employees
            SET full_name = :full_name,
                gender = COALESCE(:gender, gender),
                birth_place = COALESCE(:birth_place, birth_place),
                birth_date = COALESCE(:birth_date, birth_date),
                religion_id = COALESCE(:religion_id, religion_id),
                education_id = COALESCE(:education_id, education_id),
                marital_status_id = COALESCE(:marital_status_id, marital_status_id),
                is_active = :is_active,
                updated_at = now()
            WHERE employee_id = :employee_id
            """
        ),
        {**payload, "employee_id": employee_id},
    )
    return employee_id


def upsert_identity(connection, employee_id: int, ktp_number: str | None) -> None:
    ktp_number = normalize_identity_number(ktp_number)
    if not ktp_number:
        return
    existing = connection.execute(
        text(
            """
            SELECT identity_id FROM employee_identities
            WHERE employee_id = :employee_id AND identity_type = 'KTP'
            """
        ),
        {"employee_id": employee_id},
    ).first()
    if existing is not None:
        connection.execute(
            text(
                """
                UPDATE employee_identities
                SET identity_number = :identity_number,
                    is_primary = true
                WHERE identity_id = :identity_id
                """
            ),
            {"identity_id": int(existing[0]), "identity_number": ktp_number},
        )
        return
    if find_employee_by_identity(connection, ktp_number) not in {None, employee_id}:
        return
    connection.execute(
        text(
            """
            INSERT INTO employee_identities (employee_id, identity_type, identity_number, is_primary)
            VALUES (:employee_id, 'KTP', :identity_number, true)
            """
        ),
        {"employee_id": employee_id, "identity_number": ktp_number},
    )


def upsert_address(connection, employee_id: int, address_text: str | None) -> None:
    if not address_text:
        return
    existing = connection.execute(
        text(
            """
            SELECT address_id FROM employee_addresses
            WHERE employee_id = :employee_id AND address_type = 'DOMISILI'
            """
        ),
        {"employee_id": employee_id},
    ).first()
    if existing is not None:
        connection.execute(
            text(
                """
                UPDATE employee_addresses
                SET address_text = :address_text,
                    is_primary = true
                WHERE address_id = :address_id
                """
            ),
            {"address_id": int(existing[0]), "address_text": address_text},
        )
        return
    connection.execute(
        text(
            """
            INSERT INTO employee_addresses (employee_id, address_type, address_text, is_primary)
            VALUES (:employee_id, 'DOMISILI', :address_text, true)
            """
        ),
        {"employee_id": employee_id, "address_text": address_text},
    )


def upsert_assignment(
    connection,
    employee_id: int,
    estate_id: int,
    division_id: int,
    position_id: int,
    category_id: int | None,
    start_date: date | None,
    notes: str | None,
) -> None:
    existing = connection.execute(
        text(
            """
            SELECT assignment_id FROM employee_assignments
            WHERE employee_id = :employee_id AND is_current = true
            ORDER BY assignment_id DESC
            LIMIT 1
            """
        ),
        {"employee_id": employee_id},
    ).first()
    payload = {
        "employee_id": employee_id,
        "estate_id": estate_id,
        "division_id": division_id,
        "position_id": position_id,
        "category_id": category_id,
        "start_date": start_date or MANPOWER_AS_OF,
        "notes": notes,
    }
    if existing is not None:
        connection.execute(
            text(
                """
                UPDATE employee_assignments
                SET estate_id = :estate_id,
                    division_id = :division_id,
                    position_id = :position_id,
                    category_id = :category_id,
                    start_date = :start_date,
                    end_date = NULL,
                    is_current = true,
                    notes = :notes,
                    updated_at = now()
                WHERE assignment_id = :assignment_id
                """
            ),
            {**payload, "assignment_id": int(existing[0])},
        )
        return
    connection.execute(
        text(
            """
            INSERT INTO employee_assignments (
                employee_id, estate_id, division_id, position_id, category_id,
                start_date, is_current, notes
            )
            VALUES (
                :employee_id, :estate_id, :division_id, :position_id, :category_id,
                :start_date, true, :notes
            )
            """
        ),
        payload,
    )


def ensure_status_history(
    connection,
    employee_id: int,
    status_id: int | None,
    effective_date: date,
    notes: str | None,
) -> None:
    if status_id is None:
        return
    updated_future = connection.execute(
        text(
            """
            UPDATE employee_status_histories
            SET effective_date = :effective_date,
                notes = :notes
            WHERE employee_id = :employee_id
              AND employment_status_id = :status_id
              AND effective_date > :effective_date
            """
        ),
        {
            "employee_id": employee_id,
            "status_id": status_id,
            "effective_date": effective_date,
            "notes": notes,
        },
    ).rowcount
    if updated_future:
        return
    existing = connection.execute(
        text(
            """
            SELECT status_history_id FROM employee_status_histories
            WHERE employee_id = :employee_id
              AND employment_status_id = :status_id
              AND effective_date = :effective_date
            """
        ),
        {"employee_id": employee_id, "status_id": status_id, "effective_date": effective_date},
    ).first()
    if existing is not None:
        return
    connection.execute(
        text(
            """
            INSERT INTO employee_status_histories (
                employee_id, employment_status_id, effective_date, notes
            )
            VALUES (:employee_id, :status_id, :effective_date, :notes)
            """
        ),
        {
            "employee_id": employee_id,
            "status_id": status_id,
            "effective_date": effective_date,
            "notes": notes,
        },
    )


def upsert_contract(
    connection,
    employee_id: int,
    employment_type_id: int | None,
    start_date: date | None,
    notes: str | None,
) -> None:
    if employment_type_id is None:
        return
    effective_start = start_date or MANPOWER_AS_OF
    existing = connection.execute(
        text(
            """
            SELECT contract_id FROM employee_contracts
            WHERE employee_id = :employee_id
              AND employment_type_id = :employment_type_id
              AND end_date IS NULL
            ORDER BY contract_id DESC
            LIMIT 1
            """
        ),
        {"employee_id": employee_id, "employment_type_id": employment_type_id},
    ).first()
    if existing is not None:
        connection.execute(
            text(
                """
                UPDATE employee_contracts
                SET start_date = :effective_start,
                    notes = :notes,
                    updated_at = now()
                WHERE contract_id = :contract_id
                  AND (start_date <> :effective_start OR notes IS DISTINCT FROM :notes)
                """
            ),
            {
                "contract_id": int(existing[0]),
                "effective_start": effective_start,
                "notes": notes,
            },
        )
        connection.execute(
            text(
                """
                UPDATE employee_contracts
                SET end_date = :effective_start,
                    updated_at = now()
                WHERE employee_id = :employee_id
                  AND end_date IS NULL
                  AND contract_id <> :contract_id
                """
            ),
            {
                "employee_id": employee_id,
                "contract_id": int(existing[0]),
                "effective_start": effective_start,
            },
        )
        return
    connection.execute(
        text(
            """
            UPDATE employee_contracts
            SET end_date = :effective_start,
                updated_at = now()
            WHERE employee_id = :employee_id
              AND end_date IS NULL
            """
        ),
        {"employee_id": employee_id, "effective_start": effective_start},
    )
    connection.execute(
        text(
            """
            INSERT INTO employee_contracts (
                employee_id, employment_type_id, contract_no, start_date, salary_type, notes
            )
            VALUES (
                :employee_id, :employment_type_id, :contract_no, :start_date, :salary_type, :notes
            )
            """
        ),
        {
            "employee_id": employee_id,
            "employment_type_id": employment_type_id,
            "contract_no": f"IMP-{employee_id}-{employment_type_id}",
            "start_date": effective_start,
            "salary_type": "MVP",
            "notes": notes,
        },
    )


def upsert_bpjs(connection, employee_id: int, bpjs_type_id: int | None, number: str | None) -> None:
    number = normalize_card_number(number)
    if bpjs_type_id is None or not number:
        return
    existing = connection.execute(
        text(
            """
            SELECT employee_bpjs_id FROM employee_bpjs
            WHERE employee_id = :employee_id AND bpjs_type_id = :bpjs_type_id
            """
        ),
        {"employee_id": employee_id, "bpjs_type_id": bpjs_type_id},
    ).first()
    if existing is not None:
        connection.execute(
            text(
                """
                UPDATE employee_bpjs
                SET bpjs_number = :bpjs_number,
                    active_status = true,
                    updated_at = now()
                WHERE employee_bpjs_id = :employee_bpjs_id
                """
            ),
            {"employee_bpjs_id": int(existing[0]), "bpjs_number": number},
        )
        return
    connection.execute(
        text(
            """
            INSERT INTO employee_bpjs (
                employee_id, bpjs_type_id, bpjs_number, active_status, registered_date
            )
            VALUES (:employee_id, :bpjs_type_id, :bpjs_number, true, :registered_date)
            """
        ),
        {
            "employee_id": employee_id,
            "bpjs_type_id": bpjs_type_id,
            "bpjs_number": number,
            "registered_date": MANPOWER_AS_OF,
        },
    )


def import_employees(employees: Iterable[ImportedEmployee]) -> dict[str, int]:
    if engine is None:
        raise RuntimeError("Database engine belum siap. Cek konfigurasi .env.")

    stats = {"seen": 0, "employees": 0, "identities": 0, "addresses": 0, "assignments": 0}
    with engine.begin() as connection:
        company_id = ensure_company(connection)
        estate_id = ensure_estate(connection, company_id)
        bpjs_health_id = get_bpjs_type_id(connection, "KESEHATAN")
        bpjs_employment_id = get_bpjs_type_id(connection, "KETENAGAKERJAAN")

        for employee in employees:
            stats["seen"] += 1
            category_id = get_reference_id(
                connection,
                "employee_categories",
                "id",
                employee.category_code,
            )
            employment_status_id = get_reference_id(
                connection,
                "employment_statuses",
                "id",
                employee.employment_status_code,
            )
            employment_type_id = get_reference_id(
                connection,
                "employment_types",
                "id",
                employee.employment_type_code,
            )
            division_id = ensure_division(connection, estate_id, employee.division_name)
            position_id = ensure_position(connection, employee.position_name)
            employee_id = upsert_employee(connection, employee)
            stats["employees"] += 1
            note = import_note(employee)
            effective_date = effective_date_for_employee(employee)

            if employee.ktp_number:
                upsert_identity(connection, employee_id, employee.ktp_number)
                stats["identities"] += 1
            if employee.address_text:
                upsert_address(connection, employee_id, employee.address_text)
                stats["addresses"] += 1

            upsert_assignment(
                connection,
                employee_id,
                estate_id,
                division_id,
                position_id,
                category_id,
                effective_date,
                note,
            )
            stats["assignments"] += 1
            ensure_status_history(
                connection,
                employee_id,
                employment_status_id,
                effective_date,
                note,
            )
            upsert_contract(
                connection,
                employee_id,
                employment_type_id,
                effective_date,
                note,
            )
            upsert_bpjs(connection, employee_id, bpjs_health_id, employee.bpjs_health_number)
            upsert_bpjs(connection, employee_id, bpjs_employment_id, employee.bpjs_employment_number)

    return stats


def main() -> None:
    if engine is None:
        raise RuntimeError("Database engine belum siap. Cek konfigurasi .env.")

    ensure_governance_tables_with_admin()

    source_employees = [
        *read_staff_employees(),
        *read_staff_movements(),
        *read_manpower_employees(),
    ]
    employees = normalize_imported_employees(dedupe_imported_employees(source_employees))
    date_reviews = sum(1 for employee in employees if employee.notes and DATE_REVIEW_MARKER in employee.notes)
    stats = import_employees(employees)
    with engine.begin() as connection:
        batch_id = record_import_batch(
            connection,
            source_files=[MANPOWER_FILE, STAFF_FILE],
            source_rows=len(source_employees),
            unique_rows=len(employees),
            imported_rows=stats["employees"],
            date_reviews=date_reviews,
            notes="Latest HRIS manpower import from March 2026 Excel sources.",
        )
        issue_stats = sync_data_quality_issues(connection)

    print(f"source_manpower={MANPOWER_FILE.name}")
    print(f"source_staff={STAFF_FILE.name}")
    print(f"source_rows={len(source_employees)}")
    print(f"unique_rows={len(employees)}")
    print(f"date_reviews={date_reviews}")
    print(f"import_batch_id={batch_id}")
    print(f"data_quality_watch={issue_stats['watch_total']}")
    print(f"data_quality_open_info={issue_stats['open_info']}")
    for key, value in stats.items():
        print(f"{key}={value}")


if __name__ == "__main__":
    main()
